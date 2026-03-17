#!/usr/bin/env python
# coding: utf-8

"""
股票连板复盘工具
功能：获取每日涨停股票数据，按连板数和题材分类，输出到Excel
"""

import requests
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from collections import defaultdict
import re
import os


class StockReviewTool:
    def __init__(self):
        self.headers = {
            'Accept': '*/*',
            'User-Agent': 'lhb/5.20.9 (com.kaipanla.www; build:1; iOS 18.6.2) Alamofire/4.9.1',
        }
        self.base_url = 'https://apphwshhq.longhuvip.com/w1/api/index.php'

        # 预定义颜色列表（用于不同题材）- 扩展到100种颜色
        self.colors = [
            # 第一组：原有15种颜色
            'FFB6C1',  # 浅粉色
            'FFD700',  # 金色
            '87CEEB',  # 天蓝色
            '98FB98',  # 浅绿色
            'DDA0DD',  # 梅红色
            'F0E68C',  # 卡其色
            'FFB347',  # 浅橙色
            'B0E0E6',  # 粉蓝色
            'FFDAB9',  # 桃色
            'E6E6FA',  # 薰衣草色
            'F5DEB3',  # 小麦色
            'FFA07A',  # 浅鲑鱼色
            '20B2AA',  # 浅海绿色
            'DEB887',  # 硬木色
            'BC8F8F',  # 玫瑰褐色

            # 第二组：新增85种颜色
            'FF69B4',  # 热粉色
            'FF1493',  # 深粉色
            'C71585',  # 中紫红色
            'DB7093',  # 古董粉红
            'FF6347',  # 番茄色
            'FF4500',  # 橙红色
            'FF8C00',  # 深橙色
            'FFA500',  # 橙色
            'FFD700',  # 金色2
            'FFFF00',  # 黄色
            'FFFFE0',  # 浅黄色
            'FFFACD',  # 柠檬绸色
            'EEE8AA',  # 灰菊黄
            'F0E68C',  # 卡其色2
            'BDB76B',  # 深卡其色
            '9ACD32',  # 黄绿色
            '7FFF00',  # 查特酒绿
            '7CFC00',  # 草坪绿
            '00FF00',  # 酸橙色
            '32CD32',  # 酸橙绿
            '00FA9A',  # 中春绿色
            '00FF7F',  # 春绿色
            '3CB371',  # 中海绿色
            '2E8B57',  # 海绿色
            '228B22',  # 森林绿
            '008000',  # 绿色
            '006400',  # 深绿色
            '9AFF9A',  # 浅绿色2
            '90EE90',  # 浅绿色3
            '8FBC8F',  # 深海绿色
            '66CDAA',  # 中碧绿色
            '5F9EA0',  # 军服蓝
            '48D1CC',  # 中绿松石色
            '40E0D0',  # 绿松石色
            '00CED1',  # 深绿松石色
            '00FFFF',  # 青色
            'E0FFFF',  # 浅青色
            'AFEEEE',  # 灰绿松石色
            '7FFFD4',  # 碧绿色
            'B0C4DE',  # 浅钢蓝色
            'ADD8E6',  # 浅蓝色
            '87CEFA',  # 浅天蓝色
            '6495ED',  # 矢车菊蓝
            '4169E1',  # 皇家蓝
            '0000FF',  # 蓝色
            '0000CD',  # 中蓝色
            '00008B',  # 深蓝色
            '000080',  # 海军蓝
            '191970',  # 午夜蓝
            '7B68EE',  # 中板岩蓝
            '6A5ACD',  # 板岩蓝
            '483D8B',  # 深板岩蓝
            '9370DB',  # 中紫色
            '8B008B',  # 深洋红色
            '9400D3',  # 深紫罗兰色
            '9932CC',  # 深兰花紫
            'BA55D3',  # 中兰花紫
            'DA70D6',  # 兰花紫
            'EE82EE',  # 紫罗兰
            'FF00FF',  # 品红
            'FF00FF',  # 洋红
            'D8BFD8',  # 蓟色
            'DDA0DD',  # 梅红色2
            'FFE4E1',  # 薄雾玫瑰
            'FFF0F5',  # 淡紫红
            'FAEBD7',  # 古董白
            'FAF0E6',  # 亚麻色
            'FFF5EE',  # 海贝色
            'F5F5DC',  # 米色
            'FFE4C4',  # 陶坯黄
            'FFEBCD',  # 白杏仁
            'FFE4B5',  # 鹿皮鞋色
            'FFDEAD',  # 纳瓦霍白
            'F4A460',  # 沙褐色
            'D2691E',  # 巧克力色
            'CD853F',  # 秘鲁色
            'DEB887',  # 硬木色2
            'D2B48C',  # 棕褐色
            'BC8F8F',  # 玫瑰褐色2
        ]

        # 题材颜色映射文件
        self.color_map_file = 'topic_colors.json'
        self.topic_colors = self.load_topic_colors()
        self.color_index = len(self.topic_colors)

        # 每日数据存储目录
        self.data_dir = 'daily_data'
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)

        # 进阶追踪表文件
        self.progress_excel = 'stock_progress_tracker.xlsx'

    def load_topic_colors(self):
        """从文件加载题材颜色映射"""
        if os.path.exists(self.color_map_file):
            try:
                with open(self.color_map_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return {}
        return {}

    def save_topic_colors(self):
        """保存题材颜色映射到文件"""
        with open(self.color_map_file, 'w', encoding='utf-8') as f:
            json.dump(self.topic_colors, f, ensure_ascii=False, indent=2)

    def get_daily_limit_index(self, day):
        """获取各板数量统计"""
        params = {
            'Day': day,
            'PhoneOSNew': '2',
            'VerSion': '5.20.0.9',
            'a': 'DailyLimitIndex',
            'apiv': 'w41',
            'c': 'HomeDingPan'
        }
        
        try:
            response = requests.get(self.base_url, params=params, headers=self.headers)
            data = response.json()
            if data.get('errcode') == '0':
                return data.get('info', [])
            return []
        except Exception as e:
            print(f"获取板数统计失败: {e}")
            return []
    
    def get_stocks_by_pidtype(self, day, pidtype):
        """根据PidType获取股票数据
        PidType: 1=首板, 2=2板, 3=3板, 4=4板, 5=5板及以上
        返回: [(stock_data, board_num), ...]
        """
        params = {
            'Day': day,
            'Index': '0',
            'Order': '1',
            'PhoneOSNew': '2',
            'PidType': str(pidtype),
            'Type': '20',
            'VerSion': '5.20.0.9',
            'a': 'DailyLimitPerformance',
            'apiv': 'w41',
            'c': 'HomeDingPan',
            'st': '1000'
        }

        try:
            response = requests.get(self.base_url, params=params, headers=self.headers)
            data = response.json()
            if data.get('errcode') == '0' and 'info' in data:
                if len(data['info']) > 0 and isinstance(data['info'][0], list):
                    stocks = data['info'][0]
                    # 为每只股票添加板数信息
                    # PidType=1-4 直接对应板数，PidType=5 需要从连板信息解析
                    result = []
                    for stock in stocks:
                        if pidtype <= 4:
                            # PidType 1-4 直接对应板数
                            result.append((stock, pidtype))
                        else:
                            # PidType=5 是高连板，需要从stock[18]解析具体板数
                            board_info = stock[18] if len(stock) > 18 else ''
                            board_num = self.parse_board_number(board_info)
                            result.append((stock, board_num))
                    return result
                return []
            return []
        except Exception as e:
            print(f"获取PidType={pidtype}股票失败: {e}")
            return []

    def get_all_board_stocks(self, day):
        """获取所有板数的股票（1-5板及以上）
        返回: [(stock_data, board_num), ...]
        """
        all_stocks = []

        # 获取1-5板的数据（PidType=1,2,3,4,5）
        for pidtype in range(1, 6):
            stocks = self.get_stocks_by_pidtype(day, pidtype)
            if stocks:
                print(f"  PidType={pidtype}: 获取到 {len(stocks)} 只股票")
                all_stocks.extend(stocks)

        return all_stocks
    
    def parse_board_number(self, board_info):
        """解析连板信息，返回板数"""
        if not board_info or board_info == '':
            return 1  # 没有连板信息，默认为一板

        # 匹配 "X连板"、"X天Y板" 或 "Y板" 格式
        # 优先匹配 "X连板"
        match = re.search(r'(\d+)连板', board_info)
        if match:
            return int(match.group(1))

        # 其次匹配 "X天Y板" 中的 Y
        match = re.search(r'(\d+)天(\d+)板', board_info)
        if match:
            return int(match.group(2))

        # 最后匹配 "Y板"
        match = re.search(r'(\d+)板', board_info)
        if match:
            return int(match.group(1))

        return 1
    
    def classify_stocks_by_board(self, stocks_with_board):
        """按连板数分类股票
        参数: stocks_with_board - [(stock_data, board_num), ...]
        """
        board_stocks = defaultdict(list)

        for stock_data, board_num in stocks_with_board:
            # 股票数据结构：
            # [0]: 代码, [1]: 名称, [5]: 主题材, [12]: 题材列表, [18]: 连板信息
            code = stock_data[0]
            name = stock_data[1]
            main_topic = stock_data[5] if len(stock_data) > 5 else ''
            topics = stock_data[12] if len(stock_data) > 12 else ''
            board_info = stock_data[18] if len(stock_data) > 18 else ''

            # 如果board_info为空，使用传入的board_num构造
            if not board_info:
                if board_num == 1:
                    board_info = '首板'
                else:
                    board_info = f'{board_num}板'

            # 存储股票信息
            stock_info = {
                'code': code,
                'name': name,
                'main_topic': main_topic,
                'topics': topics,
                'board_info': board_info,
                'board_num': board_num
            }

            board_stocks[board_num].append(stock_info)

        return board_stocks

    def save_daily_data(self, day, board_stocks):
        """保存每日数据到JSON文件"""
        filename = os.path.join(self.data_dir, f'{day}.json')

        # 按题材统计每个板数的股票
        data = {
            'date': day,
            'boards': {}
        }

        for board_num, stocks in board_stocks.items():
            # 按题材分组
            topic_groups = defaultdict(list)
            for stock in stocks:
                topic = stock['main_topic'] if stock['main_topic'] else '其他'
                topic_groups[topic].append({
                    'code': stock['code'],
                    'name': stock['name']
                })

            # 保存该板数的题材数据
            data['boards'][str(board_num)] = {
                topic: stocks_list for topic, stocks_list in topic_groups.items()
            }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"每日数据已保存到: {filename}")

    def load_daily_data(self, day):
        """加载指定日期的数据"""
        filename = os.path.join(self.data_dir, f'{day}.json')
        if os.path.exists(filename):
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        return None

    def analyze_topic_progress(self, yesterday_data, today_data):
        """分析题材进阶情况（混合判断）
        返回: {board_num: [(topic, count, color, from_topic), ...]}
        from_topic: 如果题材发生变化，记录原来的题材；否则为None
        """
        if not yesterday_data or not today_data:
            return {}

        progress = defaultdict(list)

        # 遍历今天的每个板数
        for board_num_str, today_topics in today_data['boards'].items():
            board_num = int(board_num_str)

            # 只分析2板及以上
            if board_num < 2:
                continue

            # 查找昨天的(board_num-1)板数据
            yesterday_board = str(board_num - 1)
            if yesterday_board not in yesterday_data['boards']:
                continue

            yesterday_topics = yesterday_data['boards'][yesterday_board]

            # 方案1: 找出题材完全匹配的进阶（昨天在N板，今天在N+1板，题材相同）
            for topic in today_topics:
                if topic in yesterday_topics:
                    # 这个题材进阶了
                    count = len(today_topics[topic])
                    color = self.get_topic_color(topic)
                    progress[board_num].append((topic, count, color, None))

            # 方案2: 找出股票进阶但题材变化的情况
            # 建立昨天的股票代码到题材的映射
            yesterday_stock_to_topic = {}
            for y_topic, y_stocks in yesterday_topics.items():
                for y_stock in y_stocks:
                    yesterday_stock_to_topic[y_stock['code']] = y_topic

            # 检查今天的每个题材下的股票
            for topic, stocks in today_topics.items():
                # 如果这个题材已经在方案1中处理过，跳过
                if topic in yesterday_topics:
                    continue

                # 检查这个题材下的股票是否有从昨天进阶来的
                changed_stocks = []
                from_topics = set()

                for stock in stocks:
                    stock_code = stock['code']
                    if stock_code in yesterday_stock_to_topic:
                        # 这只股票昨天也在(board_num-1)板，但题材不同
                        changed_stocks.append(stock)
                        from_topics.add(yesterday_stock_to_topic[stock_code])

                if changed_stocks:
                    # 有股票进阶但题材变化了
                    count = len(changed_stocks)
                    color = self.get_topic_color(topic)
                    from_topic_str = '、'.join(sorted(from_topics))
                    progress[board_num].append((topic, count, color, from_topic_str))

        return progress
    
    def get_topic_color(self, topic):
        """获取题材对应的颜色（带冲突检测）"""
        if topic not in self.topic_colors:
            # 获取已使用的颜色
            used_colors = set(self.topic_colors.values())

            # 找到第一个未使用的颜色
            assigned_color = None
            for color in self.colors:
                if color not in used_colors:
                    assigned_color = color
                    break

            # 如果所有颜色都用完了，使用循环索引（但会有重复）
            if assigned_color is None:
                assigned_color = self.colors[self.color_index % len(self.colors)]
                print(f'警告: 颜色池已用完，题材 "{topic}" 使用重复颜色 {assigned_color}')

            self.topic_colors[topic] = assigned_color
            self.color_index += 1

        return self.topic_colors[topic]
    
    def export_to_excel(self, board_stocks, day, filename=None):
        """导出到Excel（板数列+题材列交替 + 题材说明列）"""
        if filename is None:
            filename = f'stock_review_{day}.xlsx'

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '连板复盘'

        # 获取最大板数
        max_board_num = max(board_stocks.keys()) if board_stocks else 1

        # 将超过8板的归类为"高连板"
        # 重新组织数据：1-8板各自一列，9板及以上合并为"高连板"
        reorganized_stocks = defaultdict(list)
        for board_num, stocks in board_stocks.items():
            if board_num <= 8:
                reorganized_stocks[board_num] = stocks
            else:
                # 超过8板的归入"高连板"（用9表示）
                reorganized_stocks[9].extend(stocks)

        # 确定显示的板数范围（1-8板 + 高连板）
        display_boards = []
        for i in range(1, 9):
            if i in reorganized_stocks:
                display_boards.append(i)
        if 9 in reorganized_stocks:
            display_boards.append(9)

        # 计算总列数：每个板数2列（股票信息+题材）+ 1列题材说明
        total_cols = len(display_boards) * 2 + 1

        # 设置列宽
        for col_idx in range(1, total_cols + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

        # 写入标题
        ws['A1'] = f'{day} 连板复盘'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        merge_end = openpyxl.utils.get_column_letter(total_cols) + '1'
        ws.merge_cells(f'A1:{merge_end}')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # 写入统计信息
        ws['A2'] = '统计信息：'
        ws['A2'].font = Font(size=11, bold=True)
        stats_text = '  '.join([f'{board_num}板: {len(stocks)}只'
                                for board_num, stocks in sorted(board_stocks.items())])
        ws['B2'] = stats_text
        ws['B2'].font = Font(size=10)
        merge_end_col2 = openpyxl.utils.get_column_letter(total_cols) + '2'
        ws.merge_cells(f'B2:{merge_end_col2}')

        # 写入列标题（板数列 + 题材列交替）
        row = 4
        col_idx = 1
        for board_num in display_boards:
            # 板数列标题
            cell = ws.cell(row=row, column=col_idx)
            count = len(reorganized_stocks.get(board_num, []))

            # 如果是9，显示为"高连板"
            if board_num == 9:
                cell.value = f'高连板 ({count}只)'
            else:
                cell.value = f'{board_num}板 ({count}只)'

            cell.font = Font(size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
            col_idx += 1

            # 题材列标题
            cell = ws.cell(row=row, column=col_idx)
            cell.value = '题材'
            cell.font = Font(size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
            col_idx += 1

        # 最后一列：题材说明
        cell = ws.cell(row=row, column=total_cols)
        cell.value = '题材说明'
        cell.font = Font(size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        ws.row_dimensions[row].height = 25

        # 按题材分组并写入数据
        max_row = 5

        for board_idx, board_num in enumerate(display_boards):
            if board_num not in reorganized_stocks:
                continue

            stocks = reorganized_stocks[board_num]

            # 按主题材分组
            topic_groups = defaultdict(list)
            for stock in stocks:
                topic = stock['main_topic'] if stock['main_topic'] else '其他'
                topic_groups[topic].append(stock)

            # 计算该板数的列索引（股票信息列和题材列）
            stock_col = board_idx * 2 + 1  # 股票信息列
            topic_col = stock_col + 1      # 题材列

            # 写入每个题材的股票
            current_row = 5
            for topic, topic_stocks in sorted(topic_groups.items(), key=lambda x: -len(x[1])):
                # 获取题材颜色
                color = self.get_topic_color(topic)

                for stock_idx, stock in enumerate(topic_stocks):
                    # 股票信息列
                    cell_stock = ws.cell(row=int(current_row), column=stock_col)

                    # 设置单元格内容
                    content_parts = [
                        f"【{stock['name']}】",
                        f"代码: {stock['code']}",
                        f"题材: {stock['topics']}"
                    ]
                    if stock['board_info']:
                        content_parts.append(f"连板: {stock['board_info']}")

                    cell_stock.value = '\n'.join(content_parts)

                    # 设置样式
                    cell_stock.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell_stock.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell_stock.font = Font(size=9)

                    # 如果是该题材的第一只股票，使用粗边框顶部
                    if stock_idx == 0:
                        cell_stock.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='medium'),  # 粗边框标记题材开始
                            bottom=Side(style='thin')
                        )
                    else:
                        cell_stock.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                    # 题材列
                    cell_topic = ws.cell(row=int(current_row), column=topic_col)
                    cell_topic.value = topic
                    cell_topic.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell_topic.alignment = Alignment(horizontal='center', vertical='center')
                    cell_topic.font = Font(size=10, bold=True)

                    # 题材列边框
                    if stock_idx == 0:
                        cell_topic.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='medium'),
                            bottom=Side(style='thin')
                        )
                    else:
                        cell_topic.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                    # 设置行高
                    ws.row_dimensions[int(current_row)].height = 70

                    current_row += 1

                # 更新最大行数
                max_row = max(max_row, int(current_row))

        # 在最右边的题材说明列填充
        summary_col = total_cols
        current_row = 5

        # 收集所有题材及其股票数量（使用原始的board_stocks，包含所有板数）
        topic_stats = defaultdict(int)
        for stocks in board_stocks.values():
            for stock in stocks:
                topic = stock['main_topic'] if stock['main_topic'] else '其他'
                topic_stats[topic] += 1

        # 按股票数量排序题材
        sorted_topics = sorted(topic_stats.items(), key=lambda x: -x[1])

        # 在题材说明列填充
        for topic, count in sorted_topics:
            color = self.get_topic_color(topic)
            cell = ws.cell(row=current_row, column=summary_col)
            cell.value = f'{topic}\n({count}只)'
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(size=10, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            ws.row_dimensions[current_row].height = 50
            current_row += 1

        # 保存颜色映射
        self.save_topic_colors()

        # 不在这里保存文件，返回workbook和filename
        return wb, filename

    def find_last_trading_day(self, day):
        """找到最近的一个交易日（有数据的日期）
        从day-1开始往前找，最多找7天
        """
        from datetime import datetime as dt, timedelta

        date_obj = dt.strptime(day, '%Y-%m-%d')

        # 往前找最多7天
        for i in range(1, 8):
            check_date = (date_obj - timedelta(days=i)).strftime('%Y-%m-%d')
            data_file = os.path.join(self.data_dir, f'{check_date}.json')

            if os.path.exists(data_file):
                print(f"找到上一个交易日: {check_date} (距今{i}天)")
                return check_date

        print(f"警告: 未找到最近7天内的交易日数据")
        return None

    def load_progress_history(self):
        """加载进阶追踪的历史数据
        优先从 stock_progress_tracker.xlsx 读取
        如果不存在，返回空列表
        """
        history_rows = []

        if os.path.exists(self.progress_excel):
            try:
                wb = openpyxl.load_workbook(self.progress_excel)
                ws = wb.active

                # 读取所有数据行（跳过表头）
                # 检查列数，兼容旧版本（12列）和新版本（13列）
                max_col = min(ws.max_column, 13)

                for row_idx in range(2, ws.max_row + 1):
                    row_data = []
                    for col_idx in range(1, 14):  # 13列
                        if col_idx <= max_col:
                            cell = ws.cell(row=row_idx, column=col_idx)
                            row_data.append({
                                'value': cell.value,
                                'font': cell.font,
                                'fill': cell.fill,
                                'alignment': cell.alignment
                            })
                        else:
                            # 旧数据没有第13列，填充空值
                            row_data.append({
                                'value': '',
                                'font': None,
                                'fill': None,
                                'alignment': None
                            })
                    history_rows.append(row_data)

                print(f"已加载 {len(history_rows)} 条历史进阶数据")
            except Exception as e:
                print(f"加载历史数据失败: {e}")

        return history_rows

    def update_progress_tracker(self, wb, day, board_stocks, max_board):
        """在workbook中添加进阶追踪sheet"""
        from datetime import datetime as dt

        # 获取星期
        weekday_map = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
        date_obj = dt.strptime(day, '%Y-%m-%d')
        weekday = weekday_map[date_obj.weekday()]

        # 找到最近的一个交易日（而不是简单的昨天）
        last_trading_day = self.find_last_trading_day(day)
        if last_trading_day:
            yesterday_data = self.load_daily_data(last_trading_day)
        else:
            yesterday_data = None
            print("警告: 未找到上一个交易日数据，无法分析题材进阶")

        # 保存今天的数据
        today_data_dict = {
            'date': day,
            'boards': {}
        }
        for board_num, stocks in board_stocks.items():
            topic_groups = defaultdict(list)
            for stock in stocks:
                topic = stock['main_topic'] if stock['main_topic'] else '其他'
                topic_groups[topic].append({
                    'code': stock['code'],
                    'name': stock['name']
                })
            today_data_dict['boards'][str(board_num)] = {
                topic: stocks_list for topic, stocks_list in topic_groups.items()
            }

        self.save_daily_data(day, board_stocks)

        # 分析题材进阶
        progress = self.analyze_topic_progress(yesterday_data, today_data_dict)

        # 加载历史数据
        history_rows = self.load_progress_history()

        # 在workbook中创建新的sheet
        ws = wb.create_sheet(title='题材进阶追踪')

        # 写入表头
        headers = ['连板高度', '日期', '星期', '2板', '3板', '4板', '5板', '6板', '7板', '8板', '9板', '10板', '股票详情']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_idx == 13:  # 股票详情列宽一些
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 50
            else:
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

        # 写入历史数据
        next_row = 2
        for row_data in history_rows:
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col_idx)
                cell.value = cell_data['value']
                if cell_data['font']:
                    cell.font = cell_data['font'].copy()
                if cell_data['fill']:
                    cell.fill = cell_data['fill'].copy()
                if cell_data['alignment']:
                    cell.alignment = cell_data['alignment'].copy()
            ws.row_dimensions[next_row].height = 40
            next_row += 1

        # 写入当天的新数据行
        # 连板高度
        cell = ws.cell(row=next_row, column=1)
        cell.value = max_board
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 日期
        cell = ws.cell(row=next_row, column=2)
        cell.value = day
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 星期
        cell = ws.cell(row=next_row, column=3)
        cell.value = weekday
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 2-10板的进阶情况
        for board_num in range(2, 11):
            col_idx = board_num + 2  # 2板在第4列，3板在第5列...
            cell = ws.cell(row=next_row, column=col_idx)

            if board_num in progress:
                # 有进阶的题材
                topics_info = progress[board_num]

                # 构建显示文本
                text_parts = []
                for item in topics_info:
                    if len(item) == 4:
                        topic, count, color, from_topic = item
                        if from_topic:
                            # 题材发生了变化
                            text_parts.append(f'{topic}({count})←{from_topic}')
                        else:
                            # 题材没有变化
                            text_parts.append(f'{topic}({count})')
                    else:
                        # 兼容旧格式（3个元素）
                        topic, count, color = item
                        text_parts.append(f'{topic}({count})')

                cell.value = '\n'.join(text_parts)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # 如果只有一个题材，用背景色
                if len(topics_info) == 1:
                    item = topics_info[0]
                    if len(item) == 4:
                        topic, count, color, from_topic = item
                    else:
                        topic, count, color = item
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(size=10, bold=True)
                else:
                    # 多个题材，用文字颜色（这里简化处理，用第一个题材的颜色）
                    # Excel单元格不支持多种文字颜色，所以用加粗表示
                    cell.font = Font(size=10, bold=True, color='FF0000')
            else:
                # 没有进阶
                cell.value = f'{board_num}板'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # 添加股票详情列（第13列）
        cell = ws.cell(row=next_row, column=13)
        stock_details = []
        for board_num in sorted(board_stocks.keys()):
            stocks = board_stocks[board_num]
            stock_list = []
            for stock in stocks:
                stock_list.append(f"{stock['name']}:{stock['code']}")
            stock_details.append(f"{board_num}板（{' ，'.join(stock_list)}）")

        cell.value = '\n'.join(stock_details)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.font = Font(size=9)

        # 设置行高（增加高度以容纳更多内容）
        ws.row_dimensions[next_row].height = 100

        # 同时更新独立的进阶追踪文件（用于下次读取历史数据）
        if os.path.exists(self.progress_excel):
            tracker_wb = openpyxl.load_workbook(self.progress_excel)
            tracker_ws = tracker_wb.active
            tracker_next_row = tracker_ws.max_row + 1
        else:
            tracker_wb = openpyxl.Workbook()
            tracker_ws = tracker_wb.active
            tracker_ws.title = '题材进阶追踪'

            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                cell = tracker_ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(size=12, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx == 13:  # 股票详情列宽一些
                    tracker_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 50
                else:
                    tracker_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

            tracker_next_row = 2

        # 复制当天数据到独立文件（包括第13列）
        for col_idx in range(1, 14):  # 改为14，包含第13列
            src_cell = ws.cell(row=next_row, column=col_idx)
            dst_cell = tracker_ws.cell(row=tracker_next_row, column=col_idx)
            dst_cell.value = src_cell.value
            if src_cell.font:
                dst_cell.font = src_cell.font.copy()
            if src_cell.fill:
                dst_cell.fill = src_cell.fill.copy()
            if src_cell.alignment:
                dst_cell.alignment = src_cell.alignment.copy()

        tracker_ws.row_dimensions[tracker_next_row].height = 100
        tracker_wb.save(self.progress_excel)

        print(f"进阶追踪sheet已添加到Excel")

    def run(self, day=None):
        """运行复盘工具"""
        if day is None:
            day = datetime.now().strftime('%Y-%m-%d')

        print(f"正在获取 {day} 的涨停数据...")

        # 获取板数统计
        board_index = self.get_daily_limit_index(day)
        if board_index:
            print(f"板数统计: 一板{board_index[0]}个, 二板{board_index[1]}个, 三板{board_index[2]}个, 四板{board_index[3]}个, 更高{board_index[4]}个")

        # 获取所有板数的股票（1-5板及以上）
        all_stocks = self.get_all_board_stocks(day)
        print(f"共获取到 {len(all_stocks)} 只涨停股票")

        if not all_stocks:
            print("没有获取到涨停股票数据")
            return

        # 按连板数分类
        board_stocks = self.classify_stocks_by_board(all_stocks)

        # 打印分类结果
        for board_num in sorted(board_stocks.keys()):
            print(f"{board_num}板: {len(board_stocks[board_num])}只")

        # 获取最高板数
        max_board = max(board_stocks.keys()) if board_stocks else 1

        # 导出到Excel（获取workbook对象）
        wb, filename = self.export_to_excel(board_stocks, day)

        # 在同一个Excel中添加进阶追踪sheet
        print(f"\n正在添加进阶追踪sheet...")
        self.update_progress_tracker(wb, day, board_stocks, max_board)

        # 保存Excel文件
        wb.save(filename)
        print(f"\n复盘完成！文件已保存: {filename}")
        print(f"  - Sheet1: 连板复盘（当天数据）")
        print(f"  - Sheet2: 题材进阶追踪（历史累积）")

        return filename


if __name__ == '__main__':
    import sys
    
    # 可以通过命令行参数指定日期，格式：YYYY-MM-DD
    day = sys.argv[1] if len(sys.argv) > 1 else None
    
    tool = StockReviewTool()
    tool.run(day)

